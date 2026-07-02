"""
Data accuracy tests — compare compute_filtered() output against Excel ground truth.
Tests every question × every country cut, then every question × gender/age segment.
Run:  python3 test_data_accuracy.py
"""

import sys
import openpyxl
import pandas as pd
from pathlib import Path
from data_loader import (
    compute_filtered, load_funnel_by_country, get_raw_df,
    MULTI_SELECT_QUESTIONS, MULTI_COL_MAP,
)

EXCEL_PATH = Path("data/260630 Wellbeing Excel base.xlsx")
PASS = "✓"
FAIL = "✗"
WARN = "~"
TOLERANCE = 0.6   # max allowed difference in percentage points (rounding)

errors   = []
warnings = []
passed   = 0


def check(label, got, expected, n_got=None, n_exp=None, check_n=False):
    global passed
    diff = abs(got - expected)
    if diff > TOLERANCE:
        errors.append(f"  {FAIL} {label}: got {got:.1f}% (n={n_got})  expected {expected:.1f}%  Δ={diff:.1f}pp")
        return False
    if check_n and n_got is not None and n_exp is not None and abs(n_got - n_exp) > 1:
        warnings.append(f"  {WARN} {label}: pct ok ({got:.1f}%) but n differs got={n_got} expected={n_exp}")
    passed += 1
    return True


# ─────────────────────────────────────────────────────────────────────────────
# 1. Load ground truth from Excel
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "═" * 70)
print("  WELLBEING DASHBOARD — DATA ACCURACY TESTS")
print("═" * 70)

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# ── Funnel by Country ground truth ────────────────────────────────────────────
funnel_ws = wb["Funnel by Country"]
funnel_rows = list(funnel_ws.iter_rows(values_only=True))

# Column positions: Kuwait n=2, Kuwait %=3, KSA n=4, KSA %=5, UAE n=6, UAE %=7, Total n=8, Total %=9
COUNTRY_COLS_XL = {
    "All countries": (8, 9),
    "Kuwait":        (2, 3),
    "KSA":           (4, 5),
    "UAE":           (6, 7),
}

excel_funnel = {}   # {qid: {response_text: {country: {pct, n}}}}
cur_q = None

for row in funnel_rows:
    cell0 = str(row[0]).strip() if row[0] else ""
    if cell0.startswith("Q") and "—" in cell0:
        cur_q = cell0.split("—")[0].strip().split()[0]
        excel_funnel[cur_q] = {}
        continue
    if cur_q and row[1] is not None and isinstance(row[2], (int, float)):
        resp_text = cell0
        excel_funnel[cur_q][resp_text] = {}
        for ctry, (ni, pi) in COUNTRY_COLS_XL.items():
            excel_funnel[cur_q][resp_text][ctry] = {
                "n":   int(row[ni]) if isinstance(row[ni], (int, float)) else 0,
                "pct": float(row[pi]) * 100 if isinstance(row[pi], (int, float)) else 0.0,
            }

# ── Profile Breakdown ground truth ────────────────────────────────────────────
prof_ws = wb["Profile Breakdown"]
prof_rows = list(prof_ws.iter_rows(values_only=True))

# Column positions (0-indexed): Male=2, Female=3, 18-24=4, 25-34=5, 35-44=6, 45-54=7, 55-64=8
PROFILE_COLS = {
    "Male":   (2, 147),
    "Female": (3, 146),
    "18–24":  (4, 33),
    "25–34":  (5, 99),
    "35–44":  (6, 83),
    "45–54":  (7, 39),
    "55–64":  (8, 36),
}

excel_profile = {}  # {qid: {resp_text: {segment: pct}}}
cur_q = None

for row in prof_rows:
    cell0 = str(row[0]).strip() if row[0] else ""
    if cell0.startswith("Q") and "—" in cell0:
        cur_q = cell0.split("—")[0].strip().split()[0]
        excel_profile[cur_q] = {}
        continue
    if cur_q and row[1] is not None and isinstance(row[2], (int, float)):
        resp_text = cell0
        excel_profile[cur_q][resp_text] = {}
        for seg, (col_idx, n_seg) in PROFILE_COLS.items():
            val = row[col_idx]
            excel_profile[cur_q][resp_text][seg] = {
                "pct": float(val) * 100 if isinstance(val, (int, float)) else 0.0,
                "n":   n_seg,
            }

wb.close()

# ─────────────────────────────────────────────────────────────────────────────
# 2. Load dashboard data
# ─────────────────────────────────────────────────────────────────────────────

funnel_data = load_funnel_by_country()
Q_IDS = list(funnel_data.keys())

# ─────────────────────────────────────────────────────────────────────────────
# 3. Test A — Country cuts (All / Kuwait / KSA / UAE) for every question
# ─────────────────────────────────────────────────────────────────────────────

print(f"\n── TEST A: Country cuts ({'All, Kuwait, KSA, UAE'}) × {len(Q_IDS)} questions\n")

for qid in Q_IDS:
    q_data   = funnel_data[qid]
    responses = q_data["responses"]
    xl_q     = excel_funnel.get(qid, {})
    q_errors = []

    for ctry_label, mkt_arg in [
        ("All countries", ["All countries"]),
        ("Kuwait",        ["Kuwait"]),
        ("KSA",           ["KSA"]),
        ("UAE",           ["UAE"]),
    ]:
        results, n_total = compute_filtered(qid, mkt_arg, [], [], responses)

        for res in results:
            raw_text = res.get("text", "")
            # match against excel using raw_text (wrapped label might differ slightly)
            xl_resp = xl_q.get(raw_text)
            if xl_resp is None:
                # try to find by partial match
                for k in xl_q:
                    if raw_text.startswith(k[:25]) or k.startswith(raw_text[:25]):
                        xl_resp = xl_q[k]
                        break
            if xl_resp is None:
                continue

            xl_pct = xl_resp[ctry_label]["pct"]
            xl_n   = xl_resp[ctry_label]["n"]

            label = f"{qid} | {ctry_label} | {raw_text[:45]}"
            ok = check(label, res["pct"], xl_pct, res["n"], xl_n, check_n=True)
            if not ok:
                q_errors.append(label)

    status = PASS if not q_errors else f"{FAIL} ({len(q_errors)} mismatch(es))"
    print(f"  {status}  {qid} — {q_data['label'][:55]}")
    for e in q_errors:
        print(f"         {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 4. Test B — Gender & Age segments (All countries) for every question
# ─────────────────────────────────────────────────────────────────────────────

print(f"\n── TEST B: Demographic segments × {len(Q_IDS)} questions\n")

GENDER_TESTS = [
    ("Male",   ["Male"],   []),
    ("Female", ["Female"], []),
]
AGE_TESTS = [
    ("18–24", [], ["18–24"]),
    ("25–34", [], ["25–34"]),
    ("35–44", [], ["35–44"]),
    ("45–54", [], ["45–54"]),
    ("55–64", [], ["55–64"]),
]
ALL_SEG_TESTS = GENDER_TESTS + AGE_TESTS

for qid in Q_IDS:
    q_data    = funnel_data[qid]
    responses = q_data["responses"]
    xl_q      = excel_profile.get(qid, {})
    q_errors  = []

    for seg_label, gender_arg, age_arg in ALL_SEG_TESTS:
        results, n_total = compute_filtered(qid, ["All countries"], gender_arg, age_arg, responses)
        _, expected_n    = PROFILE_COLS[seg_label]

        # n must match exactly
        if n_total != expected_n:
            errors.append(f"  {FAIL} {qid} | {seg_label}: base n={n_total} expected={expected_n}")
            q_errors.append(f"n mismatch for {seg_label}")

        for res in results:
            raw_text = res.get("text", "")
            xl_resp  = xl_q.get(raw_text)
            if xl_resp is None:
                for k in xl_q:
                    if raw_text.startswith(k[:25]) or k.startswith(raw_text[:25]):
                        xl_resp = xl_q[k]
                        break
            if xl_resp is None:
                continue

            xl_pct = xl_resp[seg_label]["pct"]
            xl_n   = xl_resp[seg_label]["n"]

            label = f"{qid} | {seg_label} | {raw_text[:40]}"
            # Excel Profile Breakdown only stores segment base n, not per-response n
            # so we only verify the percentage, not the n
            ok = check(label, res["pct"], xl_pct)
            if not ok:
                q_errors.append(label)

    status = PASS if not q_errors else f"{FAIL} ({len(q_errors)} mismatch(es))"
    print(f"  {status}  {qid} — {q_data['label'][:55]}")
    for e in q_errors:
        print(f"         {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 5. Test C — Cross-tab spot-checks (country AND gender)
# ─────────────────────────────────────────────────────────────────────────────

print(f"\n── TEST C: Cross-tab spot-checks (country × gender, computed vs raw DATA)\n")

df = get_raw_df()
CTRY_CODES = {"Kuwait": 1, "KSA": 2, "UAE": 3}

spot_checks = [
    # (qid, country, gender, code_val_to_check, description)
    ("Q1", "Kuwait", "Male",   3, "Professional route"),
    ("Q1", "KSA",    "Female", 1, "Digital search"),
    ("Q2", "UAE",    "Male",   1, "Can't judge quality"),
    ("Q3", "Kuwait", "Female", 1, "First response"),
    ("Q4", "KSA",    "Male",   1, "First response"),
    ("Q5", "UAE",    "Female", 1, "First response"),
]

for qid, country, gender, code_val, desc in spot_checks:
    mkt_arg    = [country]
    gender_arg = [gender]

    # Compute from dashboard
    q_data    = funnel_data[qid]
    responses  = q_data["responses"]
    results, n_total = compute_filtered(qid, mkt_arg, gender_arg, [], responses)
    dash_res   = next((r for r in results if r["code"] == code_val or
                       str(r["code"]) == str(code_val)), None)
    dash_pct   = dash_res["pct"] if dash_res else None
    dash_n     = dash_res["n"]   if dash_res else None

    # Compute directly from raw df
    ctry_code   = CTRY_CODES[country]
    gender_code = 1 if gender == "Male" else 2
    mask = (df["ctry"] == ctry_code) & (df["S2"] == gender_code)
    sub  = df[mask]
    n_sub = len(sub)
    if qid in MULTI_SELECT_QUESTIONS:
        col = next((v for k, v in MULTI_COL_MAP.items() if
                    int(k, 36) == code_val), None)
        raw_pct = float(sub[col].sum() / n_sub) * 100 if col and n_sub else None
        raw_n   = int(sub[col].sum()) if col else None
    else:
        raw_pct = float((sub[qid] == code_val).sum() / n_sub) * 100 if n_sub else 0.0
        raw_n   = int((sub[qid] == code_val).sum())

    label = f"{qid} | {country} × {gender} | {desc}"
    if dash_pct is not None and raw_pct is not None:
        check(label, dash_pct, raw_pct, dash_n, raw_n)
        print(f"       dashboard={dash_pct:.1f}%  raw={raw_pct:.1f}%  n_sub={n_sub}")
    else:
        print(f"  ? {label}: could not compare (dash={dash_pct}, raw={raw_pct})")

# ─────────────────────────────────────────────────────────────────────────────
# 6. Summary
# ─────────────────────────────────────────────────────────────────────────────

total = passed + len(errors)
print("\n" + "═" * 70)
print(f"  RESULTS:  {passed}/{total} checks passed")
if warnings:
    print(f"\n  WARNINGS ({len(warnings)}):")
    for w in warnings:
        print(w)
if errors:
    print(f"\n  FAILURES ({len(errors)}):")
    for e in errors:
        print(e)
    print()
    sys.exit(1)
else:
    print("\n  All checks passed — dashboard data matches Excel 100%\n")
