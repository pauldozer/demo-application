"""Loads and parses survey data from the Excel workbook."""
import textwrap
import openpyxl
import pandas as pd
from pathlib import Path

EXCEL_PATH = Path(__file__).parent / "data" / "260630 Wellbeing Excel base.xlsx"

COLORS = {
    "primary":   "#1F4E79",
    "secondary": "#7A8CA5",
    "accent":    "#4E7C59",
    "light":     "#D6E4F0",
    "gray1":     "#F5F6F8",
    "gray2":     "#E8EAED",
    "gray3":     "#9AA3AE",
    "white":     "#FFFFFF",
}

# Country mapping in DATA sheet
CTRY_MAP     = {1: "Kuwait", 2: "KSA", 3: "UAE"}
CTRY_REVERSE = {"Kuwait": 1, "KSA": 2, "UAE": 3}

GENDER_MAP = {1: "Male", 2: "Female"}
AGE_MAP    = {2: "18–24", 3: "25–34", 4: "35–44", 5: "45–54", 6: "55–64", 7: "65+"}

PROFILE_SEGMENTS = {
    "Gender": {"col": "S2", "map": GENDER_MAP},
    "Age":    {"col": "S1", "map": AGE_MAP},
}

MULTI_SELECT_QUESTIONS = {"Q7", "Q9"}

# Maps letter code in Funnel sheet → DATA column name
MULTI_COL_MAP = {
    "W": "Q7_1", "X": "Q7_2", "Y": "Q7_3",
    "Z": "Q7_4", "AA": "Q7_5", "AB": "Q7_6",
    "AD": "Q9_1", "AE": "Q9_2", "AF": "Q9_3",
    "AG": "Q9_4", "AH": "Q9_5", "AI": "Q9_6",
}

COUNTRY_COLS = {
    "Kuwait": (2, 3), "KSA": (4, 5), "UAE": (6, 7), "Total": (8, 9),
}


def _open():
    return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)


def wrap_label(text: str, width: int = 38) -> str:
    """Wrap long labels with <br> for Plotly."""
    if len(text) <= width:
        return text
    lines = textwrap.wrap(text, width=width, break_long_words=False)
    return "<br>".join(lines)


# ── Funnel by Country ────────────────────────────────────────────────────────

def get_bases_funnel() -> dict:
    wb = _open()
    ws = wb["Funnel by Country"]
    for row in ws.iter_rows(max_row=10, values_only=True):
        if row[0] and "Base" in str(row[0]):
            wb.close()
            return {"Kuwait": row[2], "KSA": row[4], "UAE": row[6], "Total": row[8]}
    wb.close()
    return {}


def _qid(cell: str) -> str:
    return cell.split("—")[0].strip().split(" ")[0]


def load_funnel_by_country() -> dict:
    wb = _open()
    ws = wb["Funnel by Country"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    questions = {}
    current_q = None

    for row in rows:
        if not any(c is not None for c in row):
            continue
        cell0 = str(row[0]).strip() if row[0] is not None else ""

        if cell0.startswith("Q") and "—" in cell0:
            current_q = _qid(cell0)
            questions[current_q] = {
                "label":     cell0.split("—", 1)[1].strip(),
                "multi":     current_q in MULTI_SELECT_QUESTIONS,
                "responses": [],
            }
            continue

        if cell0 in ("Response", "Base (n):", "Total", ""):
            continue

        if current_q and row[1] is not None and isinstance(row[2], (int, float)):
            entry = {"text": wrap_label(cell0), "raw_text": cell0, "code": row[1]}
            for country, (ni, pi) in COUNTRY_COLS.items():
                entry[country] = {
                    "n":   row[ni] if isinstance(row[ni], (int, float)) else 0,
                    "pct": row[pi] if isinstance(row[pi], (int, float)) else 0.0,
                }
            questions[current_q]["responses"].append(entry)

    return questions


# ── Raw DATA sheet ────────────────────────────────────────────────────────────

_raw_df = None


def get_raw_df() -> pd.DataFrame:
    global _raw_df
    if _raw_df is None:
        wb = _open()
        ws = wb["DATA"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        df = pd.DataFrame(rows[1:], columns=rows[0])
        df = df[df["sys_RespStatus"] == 5].copy()
        _raw_df = df
    return _raw_df


# ── Cohort definitions (for v2 unified chart) ─────────────────────────────────

COHORTS = {
    "All countries": {"type": "all"},
    "Kuwait":        {"type": "country", "ctry": 1},
    "KSA":           {"type": "country", "ctry": 2},
    "UAE":           {"type": "country", "ctry": 3},
    "Male":          {"type": "demo", "col": "S2", "val": 1},
    "Female":        {"type": "demo", "col": "S2", "val": 2},
    "18–24":         {"type": "demo", "col": "S1", "val": 2},
    "25–34":         {"type": "demo", "col": "S1", "val": 3},
    "35–44":         {"type": "demo", "col": "S1", "val": 4},
    "45–54":         {"type": "demo", "col": "S1", "val": 5},
    "55–64":         {"type": "demo", "col": "S1", "val": 6},
}

COHORT_COLORS = {
    "All countries": "#1F4E79",
    "Kuwait":        "#2D7DD2",
    "KSA":           "#5BAFD6",
    "UAE":           "#97C8EB",
    "Male":          "#4E7C59",
    "Female":        "#82B494",
    "18–24":         "#C0774A",
    "25–34":         "#D4956A",
    "35–44":         "#E8B38A",
    "45–54":         "#7B5EA7",
    "55–64":         "#A990C8",
}

# Ordered for display: markets first, then gender, then age
COHORT_ORDER = [
    "All countries", "Kuwait", "KSA", "UAE",
    "Male", "Female",
    "18–24", "25–34", "35–44", "45–54", "55–64",
]

FILTER_GROUPS = {
    "Market":  ["All countries", "Kuwait", "KSA", "UAE"],
    "Gender":  ["Male", "Female"],
    "Age":     ["18–24", "25–34", "35–44", "45–54", "55–64"],
}


def compute_filtered(qid, mkt_sel, gender_sel, age_sel, response_labels):
    """
    Apply AND logic across filter groups, OR within each group.

    - mkt_sel:    list of Market selections  (e.g. ["Kuwait", "KSA"])
    - gender_sel: list of Gender selections  (e.g. ["Male"])
    - age_sel:    list of Age selections     (e.g. ["25–34", "35–44"])

    "All countries" or empty mkt_sel → no country filter.
    Empty gender_sel / age_sel → no filter on that dimension.

    Returns (results, n_total)
    where results = [{text, code, pct, n}, ...].
    """
    df = get_raw_df().copy()

    # ── Market (OR within, AND with other groups) ──────────────────────────
    specific_countries = [c for c in (mkt_sel or []) if c != "All countries"]
    if specific_countries:
        codes = [CTRY_REVERSE[c] for c in specific_countries if c in CTRY_REVERSE]
        if codes:
            df = df[df["ctry"].isin(codes)]

    # ── Gender (OR within) ─────────────────────────────────────────────────
    if gender_sel:
        g_codes = [k for k, v in GENDER_MAP.items() if v in gender_sel]
        if g_codes:
            df = df[df["S2"].isin(g_codes)]

    # ── Age (OR within) ───────────────────────────────────────────────────
    if age_sel:
        a_codes = [k for k, v in AGE_MAP.items() if v in age_sel]
        if a_codes:
            df = df[df["S1"].isin(a_codes)]

    n_total  = len(df)
    is_multi = qid in MULTI_SELECT_QUESTIONS
    results  = []

    for resp in response_labels:
        raw_code = resp["code"]
        if is_multi:
            data_col = MULTI_COL_MAP.get(str(raw_code))
            pct = float(df[data_col].sum() / n_total) if (data_col and data_col in df.columns and n_total) else 0.0
        else:
            try:
                code_val = int(raw_code)
            except (ValueError, TypeError):
                code_val = raw_code
            pct = float((df[qid] == code_val).sum() / n_total) if n_total else 0.0

        results.append({
            "text": resp["text"],
            "code": raw_code,
            "pct":  pct * 100,
            "n":    round(pct * n_total),
        })

    return results, n_total


def compute_segment_breakdown(qid: str, country: str, segment: str,
                               response_labels: list) -> list:
    """
    Compute cross-tab: for each response option, what % chose it within each segment bucket.
    Returns list of {text, Male: pct, Female: pct, ...} (keys = segment labels).
    """
    df = get_raw_df().copy()

    if country != "Total":
        df = df[df["ctry"] == CTRY_REVERSE[country]]

    seg_info = PROFILE_SEGMENTS[segment]
    seg_col  = seg_info["col"]
    seg_map  = seg_info["map"]

    # Bases per segment
    bases = {}
    for code, label in seg_map.items():
        bases[label] = int((df[seg_col] == code).sum())

    results = []
    is_multi = qid in MULTI_SELECT_QUESTIONS

    for resp in response_labels:
        raw_code = resp["code"]
        entry    = {"text": resp["text"], "code": raw_code}

        if is_multi:
            data_col = MULTI_COL_MAP.get(str(raw_code))
            if data_col is None or data_col not in df.columns:
                for label in seg_map.values():
                    entry[label] = 0.0
            else:
                for code, label in seg_map.items():
                    mask  = df[seg_col] == code
                    total = mask.sum()
                    entry[label] = float(df.loc[mask, data_col].sum() / total) if total else 0.0
        else:
            # single-select: code is numeric
            try:
                code_val = int(raw_code)
            except (ValueError, TypeError):
                code_val = raw_code
            for scode, label in seg_map.items():
                mask  = df[seg_col] == scode
                total = mask.sum()
                entry[label] = float((df.loc[mask, qid] == code_val).sum() / total) if total else 0.0

        results.append(entry)

    return results, bases
